#!/usr/bin/env python3
"""
Nuveen Housing LP Financial Model Generator

Generates a professional Excel financial model for a Limited Partnership
investing in LIHTC loans, municipal bonds, JVs, securitization residuals
(B-Pieces), and cash. Includes leverage (repo/TOB), Bloomberg rate
integration, and full securitization/resecuritization waterfall modeling.
Styled with Nuveen corporate branding.

Usage:
    python3 generate_lp_model.py              # Use default rates
    python3 generate_lp_model.py --bloomberg   # Pull rates from Bloomberg

Output: LP_Financial_Model.xlsx
"""

import argparse
import sys

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment, NamedStyle, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

# =============================================================================
# NUVEEN BRANDING CONSTANTS (from Nuveen_PPT-Template_Universal-Print.pptx)
# =============================================================================

DARK_TEAL = "00313C"
WHITE = "FFFFFF"
LIGHT_GRAY = "BABCBC"
ORANGE = "FF8D19"
LIME_GREEN = "C3D62E"
TEAL_GREEN = "00AD97"
MEDIUM_GRAY = "747476"

INPUT_BG = "FFF3E0"       # Light orange/peach for input cells
CALC_BG = "F0F1F1"        # Very light gray for calculated cells
SUBTOTAL_BG = "E8E8E8"    # Slightly darker gray for subtotal rows

FONT_NAME = "Georgia"
NUM_YEARS = 20

ASSET_CLASSES = [
    "LIHTC Loans",
    "Municipal Bonds",
    "JVs",
    "Securitization Residuals (B-Pieces)",
    "Cash",
]
NUM_ASSETS = len(ASSET_CLASSES)

# Index positions for key asset classes (0-based)
IDX_LIHTC = 0
IDX_BONDS = 1
IDX_JV = 2
IDX_BPIECES = 3
IDX_CASH = 4

# =============================================================================
# DEFAULT INPUTS
# =============================================================================

DEFAULT_AUM = 500_000_000
DEFAULT_MGMT_FEE_BPS = 50
DEFAULT_PERF_FEE_BPS = 2000    # 20% of excess return
DEFAULT_HURDLE_BPS = 400       # 4% hurdle
DEFAULT_OTHER_EXP_BPS = 15

# Year 1 defaults (same for all years as starting point)
DEFAULT_ALLOCATIONS = [0.35, 0.25, 0.15, 0.15, 0.10]
DEFAULT_YIELDS = [0.065, 0.045, 0.080, 0.075, 0.020]

# Leverage defaults (applied only to LIHTC loans)
DEFAULT_LEVERAGE_PCT = 0.25       # 25% of LIHTC loan AUM
DEFAULT_REPO_WEIGHT = 0.60        # 60% repo / 40% TOB split

# Securitization defaults (0% = off by default)
DEFAULT_LOAN_SECURITIZE_PCT = 0.0
DEFAULT_LOAN_ADVANCE_RATE = 0.85
DEFAULT_LOAN_SECURITIZE_SPREAD_BPS = 150
DEFAULT_JV_RESECURITIZE_PCT = 0.0
DEFAULT_JV_ADVANCE_RATE = 0.75
DEFAULT_JV_RESECURITIZE_SPREAD_BPS = 200

# Market rate defaults (approximate forward curves)
DEFAULT_SOFR_CURVE = [
    0.0430, 0.0415, 0.0400, 0.0385, 0.0370,
    0.0355, 0.0345, 0.0335, 0.0325, 0.0320,
    0.0315, 0.0310, 0.0308, 0.0305, 0.0303,
    0.0300, 0.0300, 0.0300, 0.0300, 0.0300,
]
DEFAULT_SIFMA_CURVE = [
    0.0350, 0.0340, 0.0330, 0.0320, 0.0310,
    0.0300, 0.0293, 0.0285, 0.0280, 0.0275,
    0.0270, 0.0268, 0.0265, 0.0263, 0.0260,
    0.0258, 0.0255, 0.0253, 0.0250, 0.0250,
]
DEFAULT_REPO_SPREAD_BPS = 25
DEFAULT_TOB_SPREAD_BPS = 15

# =============================================================================
# BLOOMBERG API INTEGRATION (optional)
# =============================================================================

def fetch_bloomberg_rates():
    """Attempt to pull SOFR and SIFMA forward curves from Bloomberg.

    Requires the ``blpapi`` package and an active Bloomberg Terminal
    connection.  Returns a dict with keys ``sofr``, ``sifma`` (each a
    list of 20 annual rates).  Falls back to defaults if unavailable.
    """
    try:
        import blpapi  # type: ignore
    except ImportError:
        print("[Bloomberg] blpapi not installed -- using default rate curves.")
        print("  Install with: pip install blpapi")
        return None

    try:
        session_opts = blpapi.SessionOptions()
        session_opts.setServerHost("localhost")
        session_opts.setServerPort(8194)
        session = blpapi.Session(session_opts)
        if not session.start():
            print("[Bloomberg] Failed to start session -- using defaults.")
            return None
        if not session.openService("//blp/refdata"):
            print("[Bloomberg] Failed to open refdata service -- using defaults.")
            session.stop()
            return None

        svc = session.getService("//blp/refdata")

        def _pull_curve(ticker, num_points=NUM_YEARS):
            """Pull a forward rate curve for *ticker* via BDP."""
            req = svc.createRequest("ReferenceDataRequest")
            req.append("securities", ticker)
            req.append("fields", "PX_LAST")
            session.sendRequest(req)
            rates = []
            while True:
                ev = session.nextEvent(500)
                for msg in ev:
                    if msg.hasElement("securityData"):
                        sec = msg.getElement("securityData").getValueAsElement(0)
                        fd = sec.getElement("fieldData")
                        spot = fd.getElementAsFloat("PX_LAST") / 100.0
                        rates.append(spot)
                if ev.eventType() == blpapi.Event.RESPONSE:
                    break
            # Build a simple declining curve from the spot rate
            if rates:
                spot = rates[0]
                curve = []
                for yr in range(num_points):
                    fwd = max(spot - yr * 0.0008, spot * 0.70)
                    curve.append(round(fwd, 4))
                return curve
            return None

        sofr = _pull_curve("SOFRRATE Index")
        sifma = _pull_curve("MUNIPSA Index")
        session.stop()

        result = {}
        if sofr:
            result["sofr"] = sofr
            print(f"[Bloomberg] SOFR curve pulled: {sofr[0]:.2%} -> {sofr[-1]:.2%}")
        if sifma:
            result["sifma"] = sifma
            print(f"[Bloomberg] SIFMA curve pulled: {sifma[0]:.2%} -> {sifma[-1]:.2%}")

        return result if result else None

    except Exception as exc:
        print(f"[Bloomberg] Error: {exc} -- using default rate curves.")
        return None


# =============================================================================
# STYLE HELPERS
# =============================================================================

def make_font(bold=False, color=DARK_TEAL, size=10, italic=False):
    return Font(name=FONT_NAME, bold=bold, color=color, size=size, italic=italic)

def make_fill(color):
    return PatternFill(start_color=color, end_color=color, fill_type="solid")

def make_border(bottom=None, top=None, left=None, right=None):
    def side(style):
        return Side(style=style, color=MEDIUM_GRAY) if style else None
    return Border(
        bottom=side(bottom), top=side(top),
        left=side(left), right=side(right),
    )

HEADER_FONT = make_font(bold=True, color=WHITE, size=11)
HEADER_FILL = make_fill(DARK_TEAL)
SUBHEADER_FONT = make_font(bold=True, color=DARK_TEAL, size=10)
SUBHEADER_FILL = make_fill(LIGHT_GRAY)
BODY_FONT = make_font()
INPUT_FONT = make_font(color="1A237E")  # Dark blue for input text
INPUT_FILL = make_fill(INPUT_BG)
CALC_FILL = make_fill(CALC_BG)
SUBTOTAL_FILL = make_fill(SUBTOTAL_BG)
BOLD_FONT = make_font(bold=True)
TITLE_FONT = make_font(bold=True, color=WHITE, size=14)
THIN_BORDER = make_border(bottom="thin", top="thin", left="thin", right="thin")
BOTTOM_BORDER = make_border(bottom="medium")

CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")

FMT_CURRENCY = '$#,##0'
FMT_CURRENCY_DEC = '$#,##0.00'
FMT_PCT = '0.00%'
FMT_BPS = '0" bps"'
FMT_NUMBER = '#,##0'
FMT_PCT_INPUT = '0.0%'

# =============================================================================
# CELL STYLING UTILITIES
# =============================================================================

def style_header_row(ws, row, start_col, end_col, merge=False):
    """Apply dark teal header styling to a row."""
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

def style_subheader_row(ws, row, start_col, end_col):
    """Apply subheader styling (light gray background, bold teal text)."""
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.border = THIN_BORDER

def style_input_cell(cell, fmt=None):
    """Style a cell as an editable input."""
    cell.font = INPUT_FONT
    cell.fill = INPUT_FILL
    cell.border = THIN_BORDER
    cell.alignment = RIGHT
    if fmt:
        cell.number_format = fmt

def style_calc_cell(cell, fmt=None, bold=False):
    """Style a cell as a calculated formula cell."""
    cell.font = BOLD_FONT if bold else BODY_FONT
    cell.fill = CALC_FILL
    cell.border = THIN_BORDER
    cell.alignment = RIGHT
    if fmt:
        cell.number_format = fmt

def style_subtotal_cell(cell, fmt=None):
    """Style a subtotal/total row cell."""
    cell.font = BOLD_FONT
    cell.fill = SUBTOTAL_FILL
    cell.border = Border(
        bottom=Side(style="medium", color=DARK_TEAL),
        top=Side(style="thin", color=MEDIUM_GRAY),
        left=Side(style="thin", color=MEDIUM_GRAY),
        right=Side(style="thin", color=MEDIUM_GRAY),
    )
    cell.alignment = RIGHT
    if fmt:
        cell.number_format = fmt

def style_label_cell(cell, bold=False, indent=0):
    """Style a row label cell."""
    cell.font = BOLD_FONT if bold else BODY_FONT
    cell.alignment = Alignment(
        horizontal="left", vertical="center", indent=indent
    )
    cell.border = THIN_BORDER


# =============================================================================
# MARKET RATES ROW CONSTANTS (used by Annual Model cross-references)
# =============================================================================

MR_ROW_SOFR = 5
MR_ROW_SIFMA = 6
MR_ROW_REPO_SPREAD = 9
MR_ROW_TOB_SPREAD = 10
MR_ROW_REPO_ALLIN = 13
MR_ROW_TOB_ALLIN = 14


# =============================================================================
# SHEET: MARKET RATES
# =============================================================================

def build_market_rates_sheet(wb, sofr_curve=None, sifma_curve=None):
    """Build the Market Rate Assumptions sheet with SOFR/SIFMA curves."""
    ws = wb.create_sheet("Market Rates", 1)  # Insert after Summary
    ws.sheet_properties.tabColor = LIME_GREEN

    sofr = sofr_curve or DEFAULT_SOFR_CURVE
    sifma = sifma_curve or DEFAULT_SIFMA_CURVE

    # Column widths
    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 8
    for yr in range(1, NUM_YEARS + 1):
        ws.column_dimensions[get_column_letter(2 + yr)].width = 14

    first_col = 3
    last_col = 2 + NUM_YEARS

    # ── Title Row ──
    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    title = ws.cell(row=row, column=1,
                    value="Market Rate Assumptions (Bloomberg)")
    title.font = TITLE_FONT
    title.fill = HEADER_FILL
    title.alignment = CENTER
    for c in range(1, last_col + 1):
        ws.cell(row=row, column=c).fill = HEADER_FILL

    # ── Benchmark Rates Section ──
    row = 3
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    ws.cell(row=row, column=1, value="BENCHMARK RATES").font = HEADER_FONT
    style_header_row(ws, row, 1, last_col)

    # Year labels (row 4)
    row = 4
    ws.cell(row=row, column=1).fill = SUBHEADER_FILL
    ws.cell(row=row, column=1).border = THIN_BORDER
    ws.cell(row=row, column=2).fill = SUBHEADER_FILL
    ws.cell(row=row, column=2).border = THIN_BORDER
    for yr in range(1, NUM_YEARS + 1):
        col = 2 + yr
        cell = ws.cell(row=row, column=col, value=f"Year {yr}")
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # SOFR (row 5)
    row = MR_ROW_SOFR
    label = ws.cell(row=row, column=1,
                    value="SOFR (Secured Overnight Financing Rate)")
    style_label_cell(label, indent=1)
    ws.cell(row=row, column=2, value="%").font = make_font(color=MEDIUM_GRAY, size=9)
    ws.cell(row=row, column=2).border = THIN_BORDER
    for yr in range(1, NUM_YEARS + 1):
        cell = ws.cell(row=row, column=2 + yr, value=sofr[yr - 1])
        style_input_cell(cell, FMT_PCT)

    # SIFMA (row 6)
    row = MR_ROW_SIFMA
    label = ws.cell(row=row, column=1,
                    value="SIFMA (Municipal Swap Index)")
    style_label_cell(label, indent=1)
    ws.cell(row=row, column=2, value="%").font = make_font(color=MEDIUM_GRAY, size=9)
    ws.cell(row=row, column=2).border = THIN_BORDER
    for yr in range(1, NUM_YEARS + 1):
        cell = ws.cell(row=row, column=2 + yr, value=sifma[yr - 1])
        style_input_cell(cell, FMT_PCT)

    # ── Spread Assumptions Section ──
    row = 8
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    ws.cell(row=row, column=1, value="SPREAD ASSUMPTIONS").font = HEADER_FONT
    style_header_row(ws, row, 1, last_col)

    # Repo Spread (row 9)
    row = MR_ROW_REPO_SPREAD
    label = ws.cell(row=row, column=1, value="Repo Spread over SOFR")
    style_label_cell(label, indent=1)
    ws.cell(row=row, column=2, value="bps").font = make_font(color=MEDIUM_GRAY, size=9)
    ws.cell(row=row, column=2).border = THIN_BORDER
    for yr in range(1, NUM_YEARS + 1):
        cell = ws.cell(row=row, column=2 + yr, value=DEFAULT_REPO_SPREAD_BPS)
        style_input_cell(cell, FMT_BPS)

    # TOB Spread (row 10)
    row = MR_ROW_TOB_SPREAD
    label = ws.cell(row=row, column=1, value="TOB Spread over SIFMA")
    style_label_cell(label, indent=1)
    ws.cell(row=row, column=2, value="bps").font = make_font(color=MEDIUM_GRAY, size=9)
    ws.cell(row=row, column=2).border = THIN_BORDER
    for yr in range(1, NUM_YEARS + 1):
        cell = ws.cell(row=row, column=2 + yr, value=DEFAULT_TOB_SPREAD_BPS)
        style_input_cell(cell, FMT_BPS)

    # ── All-In Financing Rates (Calculated) ──
    row = 12
    ws.cell(row=row, column=1, value="ALL-IN FINANCING RATES").font = SUBHEADER_FONT
    style_subheader_row(ws, row, 1, last_col)
    ws.cell(row=row, column=1).alignment = LEFT

    # Repo All-In (row 13)
    row = MR_ROW_REPO_ALLIN
    label = ws.cell(row=row, column=1, value="Repo All-In Rate (SOFR + Spread)")
    style_label_cell(label, indent=1)
    ws.cell(row=row, column=2, value="%").font = make_font(color=MEDIUM_GRAY, size=9)
    ws.cell(row=row, column=2).border = THIN_BORDER
    for yr in range(1, NUM_YEARS + 1):
        col = 2 + yr
        cl = get_column_letter(col)
        cell = ws.cell(row=row, column=col)
        cell.value = f"={cl}{MR_ROW_SOFR}+{cl}{MR_ROW_REPO_SPREAD}/10000"
        style_calc_cell(cell, FMT_PCT, bold=True)

    # TOB All-In (row 14)
    row = MR_ROW_TOB_ALLIN
    label = ws.cell(row=row, column=1, value="TOB All-In Rate (SIFMA + Spread)")
    style_label_cell(label, indent=1)
    ws.cell(row=row, column=2, value="%").font = make_font(color=MEDIUM_GRAY, size=9)
    ws.cell(row=row, column=2).border = THIN_BORDER
    for yr in range(1, NUM_YEARS + 1):
        col = 2 + yr
        cl = get_column_letter(col)
        cell = ws.cell(row=row, column=col)
        cell.value = f"={cl}{MR_ROW_SIFMA}+{cl}{MR_ROW_TOB_SPREAD}/10000"
        style_calc_cell(cell, FMT_PCT, bold=True)

    # ── Legend ──
    row = 16
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    legend = ws.cell(row=row, column=1,
                     value="Enter forward rates from Bloomberg Terminal. "
                           "Use SOFRRATE Index (SOFR) and MUNIPSA Index (SIFMA).")
    legend.font = make_font(italic=True, color=MEDIUM_GRAY, size=9)

    swatch = ws.cell(row=row, column=7)
    swatch.fill = INPUT_FILL
    swatch.value = "= Input"
    swatch.font = make_font(size=9, color="1A237E")
    swatch.border = THIN_BORDER

    ws.freeze_panes = "C5"
    return ws


# =============================================================================
# SHEET: ANNUAL MODEL
# =============================================================================

def build_annual_model_sheet(wb):
    """Build the Annual Model sheet with leverage and securitization."""
    ws = wb.create_sheet("Annual Model")
    ws.sheet_properties.tabColor = ORANGE

    # Column widths
    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 8
    for yr in range(1, NUM_YEARS + 1):
        col_letter = get_column_letter(2 + yr)
        ws.column_dimensions[col_letter].width = 16

    first_col = 3  # Column C = Year 1
    last_col = 2 + NUM_YEARS
    last_col_letter = get_column_letter(last_col)

    # ── Title Row ──
    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=last_col)
    title = ws.cell(row=r, column=1, value="Annual Financial Model")
    title.font = TITLE_FONT
    title.fill = HEADER_FILL
    title.alignment = CENTER
    for c in range(1, last_col + 1):
        ws.cell(row=r, column=c).fill = HEADER_FILL

    # ── INPUTS SECTION ──
    r = 3
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=last_col)
    ws.cell(row=r, column=1, value="INPUTS").font = HEADER_FONT
    style_header_row(ws, r, 1, last_col)

    # Year labels (row 4)
    r = 4
    ws.cell(row=r, column=1, value="").font = SUBHEADER_FONT
    ws.cell(row=r, column=1).fill = SUBHEADER_FILL
    ws.cell(row=r, column=1).border = THIN_BORDER
    ws.cell(row=r, column=2).fill = SUBHEADER_FILL
    ws.cell(row=r, column=2).border = THIN_BORDER
    for yr in range(1, NUM_YEARS + 1):
        col = 2 + yr
        cell = ws.cell(row=r, column=col, value=f"Year {yr}")
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # Sub-header: Asset Allocation
    r = 5
    ws.cell(row=r, column=1, value="Asset Allocation").font = SUBHEADER_FONT
    ws.cell(row=r, column=2, value="%").font = make_font(color=MEDIUM_GRAY, size=9)
    style_subheader_row(ws, r, 1, last_col)
    ws.cell(row=r, column=1).alignment = LEFT

    # Asset allocation input rows
    ROW_ALLOC_START = 6
    for i, name in enumerate(ASSET_CLASSES):
        rr = ROW_ALLOC_START + i
        label = ws.cell(row=rr, column=1, value=name)
        style_label_cell(label, indent=1)
        ws.cell(row=rr, column=2).border = THIN_BORDER
        for yr in range(1, NUM_YEARS + 1):
            col = 2 + yr
            cell = ws.cell(row=rr, column=col, value=DEFAULT_ALLOCATIONS[i])
            style_input_cell(cell, FMT_PCT)

    # Total Allocation
    ROW_TOTAL_ALLOC = ROW_ALLOC_START + NUM_ASSETS  # 11
    r = ROW_TOTAL_ALLOC
    label = ws.cell(row=r, column=1, value="Total Allocation")
    style_label_cell(label, bold=True)
    ws.cell(row=r, column=2).border = THIN_BORDER
    for yr in range(1, NUM_YEARS + 1):
        col = 2 + yr
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col)
        cell.value = f"=SUM({cl}{ROW_ALLOC_START}:{cl}{ROW_ALLOC_START + NUM_ASSETS - 1})"
        style_subtotal_cell(cell, FMT_PCT)

    # Conditional formatting: highlight red if total != 100%
    red_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    red_font = Font(name=FONT_NAME, bold=True, color="B71C1C")
    for yr in range(1, NUM_YEARS + 1):
        col = 2 + yr
        cl = get_column_letter(col)
        cell_ref = f"{cl}{ROW_TOTAL_ALLOC}"
        ws.conditional_formatting.add(
            cell_ref,
            CellIsRule(operator="notEqual", formula=["1"], fill=red_fill, font=red_font)
        )

    # Sub-header: Expected Yields
    r = ROW_TOTAL_ALLOC + 1  # 12
    ROW_YIELD_HEADER = r
    ws.cell(row=r, column=1, value="Expected Yields").font = SUBHEADER_FONT
    ws.cell(row=r, column=2, value="%").font = make_font(color=MEDIUM_GRAY, size=9)
    style_subheader_row(ws, r, 1, last_col)
    ws.cell(row=r, column=1).alignment = LEFT

    # Yield input rows
    ROW_YIELD_START = ROW_YIELD_HEADER + 1  # 13
    for i, name in enumerate(ASSET_CLASSES):
        rr = ROW_YIELD_START + i
        label = ws.cell(row=rr, column=1, value=name)
        style_label_cell(label, indent=1)
        ws.cell(row=rr, column=2).border = THIN_BORDER
        for yr in range(1, NUM_YEARS + 1):
            col = 2 + yr
            cell = ws.cell(row=rr, column=col, value=DEFAULT_YIELDS[i])
            style_input_cell(cell, FMT_PCT)

    # ── Leverage Parameters ──
    r = ROW_YIELD_START + NUM_ASSETS + 1  # skip a row: 19
    ROW_LEV_HEADER = r
    ws.cell(row=r, column=1, value="Leverage Parameters (LIHTC Loans Only)").font = SUBHEADER_FONT
    ws.cell(row=r, column=2, value="%").font = make_font(color=MEDIUM_GRAY, size=9)
    style_subheader_row(ws, r, 1, last_col)
    ws.cell(row=r, column=1).alignment = LEFT

    r += 1
    ROW_LEV_PCT = r
    label = ws.cell(row=r, column=1, value="Leverage % of LIHTC Loans")
    style_label_cell(label, indent=1)
    ws.cell(row=r, column=2).border = THIN_BORDER
    for yr in range(1, NUM_YEARS + 1):
        cell = ws.cell(row=r, column=2 + yr, value=DEFAULT_LEVERAGE_PCT)
        style_input_cell(cell, FMT_PCT_INPUT)

    r += 1
    ROW_REPO_WEIGHT = r
    label = ws.cell(row=r, column=1, value="Repo Financing Weight (vs TOB)")
    style_label_cell(label, indent=1)
    ws.cell(row=r, column=2).border = THIN_BORDER
    for yr in range(1, NUM_YEARS + 1):
        cell = ws.cell(row=r, column=2 + yr, value=DEFAULT_REPO_WEIGHT)
        style_input_cell(cell, FMT_PCT_INPUT)

    # ── Securitization Parameters ──
    r += 2
    ROW_SEC_HEADER = r
    ws.cell(row=r, column=1, value="Securitization Parameters").font = SUBHEADER_FONT
    style_subheader_row(ws, r, 1, last_col)
    ws.cell(row=r, column=1).alignment = LEFT

    r += 1
    ROW_LOAN_SEC_PCT = r
    label = ws.cell(row=r, column=1, value="% of Loans to Securitize")
    style_label_cell(label, indent=1)
    ws.cell(row=r, column=2).border = THIN_BORDER
    for yr in range(1, NUM_YEARS + 1):
        cell = ws.cell(row=r, column=2 + yr, value=DEFAULT_LOAN_SECURITIZE_PCT)
        style_input_cell(cell, FMT_PCT_INPUT)

    r += 1
    ROW_LOAN_ADV_RATE = r
    label = ws.cell(row=r, column=1, value="Loan Advance Rate")
    style_label_cell(label, indent=1)
    ws.cell(row=r, column=2).border = THIN_BORDER
    for yr in range(1, NUM_YEARS + 1):
        cell = ws.cell(row=r, column=2 + yr, value=DEFAULT_LOAN_ADVANCE_RATE)
        style_input_cell(cell, FMT_PCT_INPUT)

    r += 1
    ROW_LOAN_SEC_SPREAD = r
    label = ws.cell(row=r, column=1, value="Loan Securitization Spread")
    style_label_cell(label, indent=1)
    ws.cell(row=r, column=2, value="bps").font = make_font(color=MEDIUM_GRAY, size=9)
    ws.cell(row=r, column=2).border = THIN_BORDER
    for yr in range(1, NUM_YEARS + 1):
        cell = ws.cell(row=r, column=2 + yr, value=DEFAULT_LOAN_SECURITIZE_SPREAD_BPS)
        style_input_cell(cell, FMT_BPS)

    r += 1
    ROW_JV_SEC_PCT = r
    label = ws.cell(row=r, column=1, value="% of JVs to Resecuritize")
    style_label_cell(label, indent=1)
    ws.cell(row=r, column=2).border = THIN_BORDER
    for yr in range(1, NUM_YEARS + 1):
        cell = ws.cell(row=r, column=2 + yr, value=DEFAULT_JV_RESECURITIZE_PCT)
        style_input_cell(cell, FMT_PCT_INPUT)

    r += 1
    ROW_JV_ADV_RATE = r
    label = ws.cell(row=r, column=1, value="JV Advance Rate")
    style_label_cell(label, indent=1)
    ws.cell(row=r, column=2).border = THIN_BORDER
    for yr in range(1, NUM_YEARS + 1):
        cell = ws.cell(row=r, column=2 + yr, value=DEFAULT_JV_ADVANCE_RATE)
        style_input_cell(cell, FMT_PCT_INPUT)

    r += 1
    ROW_JV_SEC_SPREAD = r
    label = ws.cell(row=r, column=1, value="JV Resecuritization Spread")
    style_label_cell(label, indent=1)
    ws.cell(row=r, column=2, value="bps").font = make_font(color=MEDIUM_GRAY, size=9)
    ws.cell(row=r, column=2).border = THIN_BORDER
    for yr in range(1, NUM_YEARS + 1):
        cell = ws.cell(row=r, column=2 + yr, value=DEFAULT_JV_RESECURITIZE_SPREAD_BPS)
        style_input_cell(cell, FMT_BPS)

    # ── Validation: Leverage % + Securitize % > 100% warning ──
    for yr in range(1, NUM_YEARS + 1):
        col = 2 + yr
        cl = get_column_letter(col)
        # Highlight Leverage % cell red if combined > 100%
        ws.conditional_formatting.add(
            f"{cl}{ROW_LEV_PCT}",
            CellIsRule(
                operator="greaterThan",
                formula=[f"1-{cl}{ROW_LOAN_SEC_PCT}"],
                fill=red_fill, font=red_font,
            )
        )
        # Highlight Securitize % cell red if combined > 100%
        ws.conditional_formatting.add(
            f"{cl}{ROW_LOAN_SEC_PCT}",
            CellIsRule(
                operator="greaterThan",
                formula=[f"1-{cl}{ROW_LEV_PCT}"],
                fill=red_fill, font=red_font,
            )
        )

    # ══════════════════════════════════════════════════════════════════════
    # PORTFOLIO PERFORMANCE
    # ══════════════════════════════════════════════════════════════════════

    r += 2  # blank row then header
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=last_col)
    ws.cell(row=r, column=1, value="PORTFOLIO PERFORMANCE").font = HEADER_FONT
    style_header_row(ws, r, 1, last_col)

    # Beginning AUM
    r += 1
    ROW_BEG_AUM = r
    label = ws.cell(row=r, column=1, value="Beginning AUM")
    style_label_cell(label, bold=True)
    ws.cell(row=r, column=2, value="$").font = make_font(color=MEDIUM_GRAY, size=9)
    ws.cell(row=r, column=2).border = THIN_BORDER
    # (formulas set after ROW_ENDING_AUM is known)

    # Gross Portfolio Yield
    r += 1
    ROW_GROSS_YIELD = r
    label = ws.cell(row=r, column=1, value="Gross Portfolio Yield")
    style_label_cell(label, bold=True)
    ws.cell(row=r, column=2, value="%").font = make_font(color=MEDIUM_GRAY, size=9)
    ws.cell(row=r, column=2).border = THIN_BORDER
    for yr in range(1, NUM_YEARS + 1):
        col = 2 + yr
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col)
        cell.value = (f"=SUMPRODUCT({cl}{ROW_ALLOC_START}:{cl}{ROW_ALLOC_START + NUM_ASSETS - 1},"
                      f"{cl}{ROW_YIELD_START}:{cl}{ROW_YIELD_START + NUM_ASSETS - 1})")
        style_calc_cell(cell, FMT_PCT, bold=True)

    # ── Income by Asset Class ──
    r += 2
    ws.cell(row=r, column=1, value="Income by Asset Class").font = SUBHEADER_FONT
    ws.cell(row=r, column=2, value="$").font = make_font(color=MEDIUM_GRAY, size=9)
    style_subheader_row(ws, r, 1, last_col)
    ws.cell(row=r, column=1).alignment = LEFT

    r += 1
    ROW_INCOME_START = r
    for i, name in enumerate(ASSET_CLASSES):
        rr = ROW_INCOME_START + i
        label = ws.cell(row=rr, column=1, value=name)
        style_label_cell(label, indent=1)
        ws.cell(row=rr, column=2).border = THIN_BORDER

        alloc_row = ROW_ALLOC_START + i
        yield_row = ROW_YIELD_START + i
        for yr in range(1, NUM_YEARS + 1):
            col = 2 + yr
            cl = get_column_letter(col)
            cell = ws.cell(row=rr, column=col)
            cell.value = f"={cl}{ROW_BEG_AUM}*{cl}{alloc_row}*{cl}{yield_row}"
            style_calc_cell(cell, FMT_CURRENCY)

    # Total Asset Income
    r = ROW_INCOME_START + NUM_ASSETS
    ROW_TOTAL_ASSET_INC = r
    label = ws.cell(row=r, column=1, value="Total Asset Income")
    style_label_cell(label, bold=True)
    ws.cell(row=r, column=2).border = THIN_BORDER
    for yr in range(1, NUM_YEARS + 1):
        col = 2 + yr
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col)
        cell.value = f"=SUM({cl}{ROW_INCOME_START}:{cl}{ROW_INCOME_START + NUM_ASSETS - 1})"
        style_subtotal_cell(cell, FMT_CURRENCY)

    # ══════════════════════════════════════════════════════════════════════
    # LEVERAGE (LIHTC LOANS ONLY)
    # ══════════════════════════════════════════════════════════════════════

    r += 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=last_col)
    ws.cell(row=r, column=1, value="LEVERAGE (LIHTC LOANS ONLY)").font = HEADER_FONT
    style_header_row(ws, r, 1, last_col)

    # LIHTC Loan AUM
    r += 1
    ROW_LIHTC_AUM = r
    label = ws.cell(row=r, column=1, value="LIHTC Loan AUM")
    style_label_cell(label, indent=1)
    ws.cell(row=r, column=2, value="$").font = make_font(color=MEDIUM_GRAY, size=9)
    ws.cell(row=r, column=2).border = THIN_BORDER
    lihtc_alloc_row = ROW_ALLOC_START + IDX_LIHTC
    for yr in range(1, NUM_YEARS + 1):
        col = 2 + yr
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col)
        cell.value = f"={cl}{ROW_BEG_AUM}*{cl}{lihtc_alloc_row}"
        style_calc_cell(cell, FMT_CURRENCY)

    # Leveraged Amount
    r += 1
    ROW_LEVERAGED_AMT = r
    label = ws.cell(row=r, column=1, value="Leveraged Amount")
    style_label_cell(label, indent=1)
    ws.cell(row=r, column=2).border = THIN_BORDER
    for yr in range(1, NUM_YEARS + 1):
        col = 2 + yr
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col)
        cell.value = f"={cl}{ROW_LIHTC_AUM}*{cl}{ROW_LEV_PCT}"
        style_calc_cell(cell, FMT_CURRENCY)

    # Repo Portion
    r += 1
    ROW_REPO_PORTION = r
    label = ws.cell(row=r, column=1, value="Repo Portion")
    style_label_cell(label, indent=2)
    ws.cell(row=r, column=2).border = THIN_BORDER
    for yr in range(1, NUM_YEARS + 1):
        col = 2 + yr
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col)
        cell.value = f"={cl}{ROW_LEVERAGED_AMT}*{cl}{ROW_REPO_WEIGHT}"
        style_calc_cell(cell, FMT_CURRENCY)

    # TOB Portion
    r += 1
    ROW_TOB_PORTION = r
    label = ws.cell(row=r, column=1, value="TOB Portion")
    style_label_cell(label, indent=2)
    ws.cell(row=r, column=2).border = THIN_BORDER
    for yr in range(1, NUM_YEARS + 1):
        col = 2 + yr
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col)
        cell.value = f"={cl}{ROW_LEVERAGED_AMT}*(1-{cl}{ROW_REPO_WEIGHT})"
        style_calc_cell(cell, FMT_CURRENCY)

    # Repo Financing Cost
    r += 1
    ROW_REPO_COST = r
    label = ws.cell(row=r, column=1, value="Repo Financing Cost")
    style_label_cell(label, indent=2)
    ws.cell(row=r, column=2).border = THIN_BORDER
    for yr in range(1, NUM_YEARS + 1):
        col = 2 + yr
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col)
        cell.value = f"={cl}{ROW_REPO_PORTION}*'Market Rates'!{cl}{MR_ROW_REPO_ALLIN}"
        style_calc_cell(cell, FMT_CURRENCY)

    # TOB Financing Cost
    r += 1
    ROW_TOB_COST = r
    label = ws.cell(row=r, column=1, value="TOB Financing Cost")
    style_label_cell(label, indent=2)
    ws.cell(row=r, column=2).border = THIN_BORDER
    for yr in range(1, NUM_YEARS + 1):
        col = 2 + yr
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col)
        cell.value = f"={cl}{ROW_TOB_PORTION}*'Market Rates'!{cl}{MR_ROW_TOB_ALLIN}"
        style_calc_cell(cell, FMT_CURRENCY)

    # Total Financing Cost
    r += 1
    ROW_TOTAL_FIN_COST = r
    label = ws.cell(row=r, column=1, value="Total Financing Cost")
    style_label_cell(label, indent=1)
    ws.cell(row=r, column=2).border = THIN_BORDER
    for yr in range(1, NUM_YEARS + 1):
        col = 2 + yr
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col)
        cell.value = f"={cl}{ROW_REPO_COST}+{cl}{ROW_TOB_COST}"
        style_calc_cell(cell, FMT_CURRENCY)

    # Gross Leverage Income
    r += 1
    ROW_GROSS_LEV_INC = r
    label = ws.cell(row=r, column=1, value="Gross Leverage Income")
    style_label_cell(label, indent=1)
    ws.cell(row=r, column=2).border = THIN_BORDER
    lihtc_yield_row = ROW_YIELD_START + IDX_LIHTC
    for yr in range(1, NUM_YEARS + 1):
        col = 2 + yr
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col)
        cell.value = f"={cl}{ROW_LEVERAGED_AMT}*{cl}{lihtc_yield_row}"
        style_calc_cell(cell, FMT_CURRENCY)

    # Net Leverage Income
    r += 1
    ROW_NET_LEV_INC = r
    label = ws.cell(row=r, column=1, value="Net Leverage Income")
    style_label_cell(label, bold=True)
    ws.cell(row=r, column=2).border = THIN_BORDER
    for yr in range(1, NUM_YEARS + 1):
        col = 2 + yr
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col)
        cell.value = f"={cl}{ROW_GROSS_LEV_INC}-{cl}{ROW_TOTAL_FIN_COST}"
        style_subtotal_cell(cell, FMT_CURRENCY)

    # ══════════════════════════════════════════════════════════════════════
    # SECURITIZATION
    # ══════════════════════════════════════════════════════════════════════

    r += 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=last_col)
    ws.cell(row=r, column=1, value="SECURITIZATION").font = HEADER_FONT
    style_header_row(ws, r, 1, last_col)

    # ── Loan Securitization Waterfall ──
    r += 1
    ws.cell(row=r, column=1, value="Loan Securitization Waterfall").font = SUBHEADER_FONT
    style_subheader_row(ws, r, 1, last_col)
    ws.cell(row=r, column=1).alignment = LEFT

    # Loan Pool for Securitization
    r += 1
    ROW_LOAN_POOL = r
    label = ws.cell(row=r, column=1, value="Loan Pool for Securitization")
    style_label_cell(label, indent=1)
    ws.cell(row=r, column=2, value="$").font = make_font(color=MEDIUM_GRAY, size=9)
    ws.cell(row=r, column=2).border = THIN_BORDER
    for yr in range(1, NUM_YEARS + 1):
        col = 2 + yr
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col)
        cell.value = f"={cl}{ROW_BEG_AUM}*{cl}{lihtc_alloc_row}*{cl}{ROW_LOAN_SEC_PCT}"
        style_calc_cell(cell, FMT_CURRENCY)

    # Senior Tranche (Proceeds)
    r += 1
    ROW_LOAN_SENIOR = r
    label = ws.cell(row=r, column=1, value="Senior Tranche (Proceeds)")
    style_label_cell(label, indent=2)
    ws.cell(row=r, column=2).border = THIN_BORDER
    for yr in range(1, NUM_YEARS + 1):
        col = 2 + yr
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col)
        cell.value = f"={cl}{ROW_LOAN_POOL}*{cl}{ROW_LOAN_ADV_RATE}"
        style_calc_cell(cell, FMT_CURRENCY)

    # Subordinate / Retained Tranche
    r += 1
    ROW_LOAN_SUB = r
    label = ws.cell(row=r, column=1, value="Subordinate / Retained Tranche")
    style_label_cell(label, indent=2)
    ws.cell(row=r, column=2).border = THIN_BORDER
    for yr in range(1, NUM_YEARS + 1):
        col = 2 + yr
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col)
        cell.value = f"={cl}{ROW_LOAN_POOL}-{cl}{ROW_LOAN_SENIOR}"
        style_calc_cell(cell, FMT_CURRENCY)

    # Gross Pool Income
    r += 1
    ROW_LOAN_GROSS_INC = r
    label = ws.cell(row=r, column=1, value="Gross Pool Income")
    style_label_cell(label, indent=2)
    ws.cell(row=r, column=2).border = THIN_BORDER
    for yr in range(1, NUM_YEARS + 1):
        col = 2 + yr
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col)
        cell.value = f"={cl}{ROW_LOAN_POOL}*{cl}{lihtc_yield_row}"
        style_calc_cell(cell, FMT_CURRENCY)

    # Senior Tranche Cost
    r += 1
    ROW_LOAN_SENIOR_COST = r
    label = ws.cell(row=r, column=1, value="Senior Tranche Cost")
    style_label_cell(label, indent=2)
    ws.cell(row=r, column=2).border = THIN_BORDER
    for yr in range(1, NUM_YEARS + 1):
        col = 2 + yr
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col)
        cell.value = f"={cl}{ROW_LOAN_SENIOR}*'Market Rates'!{cl}{MR_ROW_REPO_ALLIN}"
        style_calc_cell(cell, FMT_CURRENCY)

    # Excess Spread
    r += 1
    ROW_LOAN_EXCESS = r
    label = ws.cell(row=r, column=1, value="Excess Spread")
    style_label_cell(label, indent=2)
    ws.cell(row=r, column=2).border = THIN_BORDER
    for yr in range(1, NUM_YEARS + 1):
        col = 2 + yr
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col)
        cell.value = f"={cl}{ROW_LOAN_GROSS_INC}-{cl}{ROW_LOAN_SENIOR_COST}"
        style_calc_cell(cell, FMT_CURRENCY)

    # Issuance Cost
    r += 1
    ROW_LOAN_ISS_COST = r
    label = ws.cell(row=r, column=1, value="Issuance Cost")
    style_label_cell(label, indent=2)
    ws.cell(row=r, column=2).border = THIN_BORDER
    for yr in range(1, NUM_YEARS + 1):
        col = 2 + yr
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col)
        cell.value = f"={cl}{ROW_LOAN_POOL}*{cl}{ROW_LOAN_SEC_SPREAD}/10000"
        style_calc_cell(cell, FMT_CURRENCY)

    # Net Loan Securitization Income
    r += 1
    ROW_NET_LOAN_SEC_INC = r
    label = ws.cell(row=r, column=1, value="Net Loan Securitization Income")
    style_label_cell(label, bold=True)
    ws.cell(row=r, column=2).border = THIN_BORDER
    for yr in range(1, NUM_YEARS + 1):
        col = 2 + yr
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col)
        cell.value = f"={cl}{ROW_LOAN_EXCESS}-{cl}{ROW_LOAN_ISS_COST}"
        style_subtotal_cell(cell, FMT_CURRENCY)

    # ── JV Resecuritization Waterfall ──
    r += 1
    ws.cell(row=r, column=1, value="JV Resecuritization Waterfall").font = SUBHEADER_FONT
    style_subheader_row(ws, r, 1, last_col)
    ws.cell(row=r, column=1).alignment = LEFT

    jv_alloc_row = ROW_ALLOC_START + IDX_JV
    jv_yield_row = ROW_YIELD_START + IDX_JV

    # JV Pool for Resecuritization
    r += 1
    ROW_JV_POOL = r
    label = ws.cell(row=r, column=1, value="JV Pool for Resecuritization")
    style_label_cell(label, indent=1)
    ws.cell(row=r, column=2, value="$").font = make_font(color=MEDIUM_GRAY, size=9)
    ws.cell(row=r, column=2).border = THIN_BORDER
    for yr in range(1, NUM_YEARS + 1):
        col = 2 + yr
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col)
        cell.value = f"={cl}{ROW_BEG_AUM}*{cl}{jv_alloc_row}*{cl}{ROW_JV_SEC_PCT}"
        style_calc_cell(cell, FMT_CURRENCY)

    # Senior Tranche (Proceeds)
    r += 1
    ROW_JV_SENIOR = r
    label = ws.cell(row=r, column=1, value="Senior Tranche (Proceeds)")
    style_label_cell(label, indent=2)
    ws.cell(row=r, column=2).border = THIN_BORDER
    for yr in range(1, NUM_YEARS + 1):
        col = 2 + yr
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col)
        cell.value = f"={cl}{ROW_JV_POOL}*{cl}{ROW_JV_ADV_RATE}"
        style_calc_cell(cell, FMT_CURRENCY)

    # Subordinate / Retained Tranche
    r += 1
    ROW_JV_SUB = r
    label = ws.cell(row=r, column=1, value="Subordinate / Retained Tranche")
    style_label_cell(label, indent=2)
    ws.cell(row=r, column=2).border = THIN_BORDER
    for yr in range(1, NUM_YEARS + 1):
        col = 2 + yr
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col)
        cell.value = f"={cl}{ROW_JV_POOL}-{cl}{ROW_JV_SENIOR}"
        style_calc_cell(cell, FMT_CURRENCY)

    # Gross Pool Income
    r += 1
    ROW_JV_GROSS_INC = r
    label = ws.cell(row=r, column=1, value="Gross Pool Income")
    style_label_cell(label, indent=2)
    ws.cell(row=r, column=2).border = THIN_BORDER
    for yr in range(1, NUM_YEARS + 1):
        col = 2 + yr
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col)
        cell.value = f"={cl}{ROW_JV_POOL}*{cl}{jv_yield_row}"
        style_calc_cell(cell, FMT_CURRENCY)

    # Senior Tranche Cost
    r += 1
    ROW_JV_SENIOR_COST = r
    label = ws.cell(row=r, column=1, value="Senior Tranche Cost")
    style_label_cell(label, indent=2)
    ws.cell(row=r, column=2).border = THIN_BORDER
    for yr in range(1, NUM_YEARS + 1):
        col = 2 + yr
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col)
        cell.value = f"={cl}{ROW_JV_SENIOR}*'Market Rates'!{cl}{MR_ROW_TOB_ALLIN}"
        style_calc_cell(cell, FMT_CURRENCY)

    # Excess Spread
    r += 1
    ROW_JV_EXCESS = r
    label = ws.cell(row=r, column=1, value="Excess Spread")
    style_label_cell(label, indent=2)
    ws.cell(row=r, column=2).border = THIN_BORDER
    for yr in range(1, NUM_YEARS + 1):
        col = 2 + yr
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col)
        cell.value = f"={cl}{ROW_JV_GROSS_INC}-{cl}{ROW_JV_SENIOR_COST}"
        style_calc_cell(cell, FMT_CURRENCY)

    # Issuance Cost
    r += 1
    ROW_JV_ISS_COST = r
    label = ws.cell(row=r, column=1, value="Issuance Cost")
    style_label_cell(label, indent=2)
    ws.cell(row=r, column=2).border = THIN_BORDER
    for yr in range(1, NUM_YEARS + 1):
        col = 2 + yr
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col)
        cell.value = f"={cl}{ROW_JV_POOL}*{cl}{ROW_JV_SEC_SPREAD}/10000"
        style_calc_cell(cell, FMT_CURRENCY)

    # Net JV Resecuritization Income
    r += 1
    ROW_NET_JV_SEC_INC = r
    label = ws.cell(row=r, column=1, value="Net JV Resecuritization Income")
    style_label_cell(label, bold=True)
    ws.cell(row=r, column=2).border = THIN_BORDER
    for yr in range(1, NUM_YEARS + 1):
        col = 2 + yr
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col)
        cell.value = f"={cl}{ROW_JV_EXCESS}-{cl}{ROW_JV_ISS_COST}"
        style_subtotal_cell(cell, FMT_CURRENCY)

    # Total Net Securitization Income
    r += 1
    ROW_TOTAL_SECUR_INC = r
    label = ws.cell(row=r, column=1, value="Total Net Securitization Income")
    style_label_cell(label, bold=True)
    ws.cell(row=r, column=2).border = THIN_BORDER
    for yr in range(1, NUM_YEARS + 1):
        col = 2 + yr
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col)
        cell.value = f"={cl}{ROW_NET_LOAN_SEC_INC}+{cl}{ROW_NET_JV_SEC_INC}"
        style_subtotal_cell(cell, FMT_CURRENCY)

    # ══════════════════════════════════════════════════════════════════════
    # TOTAL GROSS INCOME
    # ══════════════════════════════════════════════════════════════════════

    r += 2
    ROW_TOTAL_GROSS = r
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    label = ws.cell(row=r, column=1, value="TOTAL GROSS INCOME")
    label.font = make_font(bold=True, color=WHITE, size=10)
    label.fill = make_fill(TEAL_GREEN)
    label.alignment = LEFT
    label.border = THIN_BORDER
    ws.cell(row=r, column=2).fill = make_fill(TEAL_GREEN)
    ws.cell(row=r, column=2).border = THIN_BORDER

    for yr in range(1, NUM_YEARS + 1):
        col = 2 + yr
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col)
        cell.value = (f"={cl}{ROW_TOTAL_ASSET_INC}"
                      f"+{cl}{ROW_NET_LEV_INC}"
                      f"+{cl}{ROW_TOTAL_SECUR_INC}")
        cell.font = make_font(bold=True, color=WHITE)
        cell.fill = make_fill(TEAL_GREEN)
        cell.number_format = FMT_CURRENCY
        cell.alignment = RIGHT
        cell.border = THIN_BORDER

    # ══════════════════════════════════════════════════════════════════════
    # FEES & EXPENSES
    # ══════════════════════════════════════════════════════════════════════

    r += 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=last_col)
    ws.cell(row=r, column=1, value="FEES & EXPENSES").font = HEADER_FONT
    style_header_row(ws, r, 1, last_col)

    # Management Fee
    r += 1
    ROW_MGMT_FEE = r
    label = ws.cell(row=r, column=1, value="Management Fee")
    style_label_cell(label, indent=1)
    ws.cell(row=r, column=2).border = THIN_BORDER
    for yr in range(1, NUM_YEARS + 1):
        col = 2 + yr
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col)
        cell.value = f"={cl}{ROW_BEG_AUM}*Summary!$C$6/10000"
        style_calc_cell(cell, FMT_CURRENCY)

    # Return Above Hurdle
    r += 1
    ROW_HURDLE = r
    label = ws.cell(row=r, column=1, value="Return Above Hurdle")
    style_label_cell(label, indent=1)
    ws.cell(row=r, column=2).border = THIN_BORDER
    for yr in range(1, NUM_YEARS + 1):
        col = 2 + yr
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col)
        cell.value = f"=MAX(0,{cl}{ROW_TOTAL_GROSS}-{cl}{ROW_BEG_AUM}*Summary!$C$8/10000)"
        style_calc_cell(cell, FMT_CURRENCY)

    # Performance Fee
    r += 1
    ROW_PERF_FEE = r
    label = ws.cell(row=r, column=1, value="Performance Fee")
    style_label_cell(label, indent=1)
    ws.cell(row=r, column=2).border = THIN_BORDER
    for yr in range(1, NUM_YEARS + 1):
        col = 2 + yr
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col)
        cell.value = f"={cl}{ROW_HURDLE}*Summary!$C$7/10000"
        style_calc_cell(cell, FMT_CURRENCY)

    # Other Expenses
    r += 1
    ROW_OTHER_EXP = r
    label = ws.cell(row=r, column=1, value="Other Expenses")
    style_label_cell(label, indent=1)
    ws.cell(row=r, column=2).border = THIN_BORDER
    for yr in range(1, NUM_YEARS + 1):
        col = 2 + yr
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col)
        cell.value = f"={cl}{ROW_BEG_AUM}*Summary!$C$9/10000"
        style_calc_cell(cell, FMT_CURRENCY)

    # Total Fees & Expenses
    r += 1
    ROW_TOTAL_FEES = r
    label = ws.cell(row=r, column=1, value="Total Fees & Expenses")
    style_label_cell(label, bold=True)
    ws.cell(row=r, column=2).border = THIN_BORDER
    for yr in range(1, NUM_YEARS + 1):
        col = 2 + yr
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col)
        cell.value = f"={cl}{ROW_MGMT_FEE}+{cl}{ROW_PERF_FEE}+{cl}{ROW_OTHER_EXP}"
        style_subtotal_cell(cell, FMT_CURRENCY)

    # ══════════════════════════════════════════════════════════════════════
    # RETURNS
    # ══════════════════════════════════════════════════════════════════════

    r += 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=last_col)
    ws.cell(row=r, column=1, value="RETURNS").font = HEADER_FONT
    style_header_row(ws, r, 1, last_col)

    # Net Income
    r += 1
    ROW_NET_INCOME = r
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    label = ws.cell(row=r, column=1, value="NET INCOME")
    label.font = make_font(bold=True, color=WHITE, size=10)
    label.fill = make_fill(DARK_TEAL)
    label.alignment = LEFT
    label.border = THIN_BORDER
    ws.cell(row=r, column=2).fill = make_fill(DARK_TEAL)
    ws.cell(row=r, column=2).border = THIN_BORDER

    for yr in range(1, NUM_YEARS + 1):
        col = 2 + yr
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col)
        cell.value = f"={cl}{ROW_TOTAL_GROSS}-{cl}{ROW_TOTAL_FEES}"
        cell.font = make_font(bold=True, color=WHITE)
        cell.fill = make_fill(DARK_TEAL)
        cell.number_format = FMT_CURRENCY
        cell.alignment = RIGHT
        cell.border = THIN_BORDER

    # Net Yield
    r += 1
    ROW_NET_YIELD = r
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    label = ws.cell(row=r, column=1, value="NET YIELD (Return to LPs)")
    label.font = make_font(bold=True, color=WHITE, size=10)
    label.fill = make_fill(DARK_TEAL)
    label.alignment = LEFT
    label.border = THIN_BORDER
    ws.cell(row=r, column=2).fill = make_fill(DARK_TEAL)
    ws.cell(row=r, column=2).border = THIN_BORDER

    for yr in range(1, NUM_YEARS + 1):
        col = 2 + yr
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col)
        cell.value = f"=IF({cl}{ROW_BEG_AUM}=0,0,{cl}{ROW_NET_INCOME}/{cl}{ROW_BEG_AUM})"
        cell.font = make_font(bold=True, color=WHITE)
        cell.fill = make_fill(DARK_TEAL)
        cell.number_format = FMT_PCT
        cell.alignment = RIGHT
        cell.border = THIN_BORDER

    # Ending AUM
    r += 1
    ROW_ENDING_AUM = r
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    label = ws.cell(row=r, column=1, value="ENDING AUM")
    label.font = make_font(bold=True, color=WHITE, size=10)
    label.fill = make_fill(ORANGE)
    label.alignment = LEFT
    label.border = THIN_BORDER
    ws.cell(row=r, column=2).fill = make_fill(ORANGE)
    ws.cell(row=r, column=2).border = THIN_BORDER

    for yr in range(1, NUM_YEARS + 1):
        col = 2 + yr
        cl = get_column_letter(col)
        cell = ws.cell(row=r, column=col)
        cell.value = f"={cl}{ROW_BEG_AUM}+{cl}{ROW_NET_INCOME}"
        cell.font = make_font(bold=True, color=WHITE)
        cell.fill = make_fill(ORANGE)
        cell.number_format = FMT_CURRENCY
        cell.alignment = RIGHT
        cell.border = THIN_BORDER

    # ── Now set Beginning AUM formulas (needs ROW_ENDING_AUM) ──
    for yr in range(1, NUM_YEARS + 1):
        col = 2 + yr
        cl = get_column_letter(col)
        cell = ws.cell(row=ROW_BEG_AUM, column=col)
        if yr == 1:
            cell.value = "=Summary!$C$5"
        else:
            prev_cl = get_column_letter(col - 1)
            cell.value = f"={prev_cl}{ROW_ENDING_AUM}"
        style_calc_cell(cell, FMT_CURRENCY, bold=True)

    # Freeze panes
    ws.freeze_panes = "C5"

    # Return row mapping for Summary sheet
    rows = {
        "beg_aum": ROW_BEG_AUM,
        "gross_yield": ROW_GROSS_YIELD,
        "total_asset_inc": ROW_TOTAL_ASSET_INC,
        "net_lev_inc": ROW_NET_LEV_INC,
        "leveraged_amt": ROW_LEVERAGED_AMT,
        "total_secur_inc": ROW_TOTAL_SECUR_INC,
        "total_gross": ROW_TOTAL_GROSS,
        "mgmt_fee": ROW_MGMT_FEE,
        "perf_fee": ROW_PERF_FEE,
        "other_exp": ROW_OTHER_EXP,
        "total_fees": ROW_TOTAL_FEES,
        "net_income": ROW_NET_INCOME,
        "net_yield": ROW_NET_YIELD,
        "ending_aum": ROW_ENDING_AUM,
    }
    return ws, rows


# =============================================================================
# SHEET: SUMMARY
# =============================================================================

def build_summary_sheet(wb, rows):
    """Build the Summary dashboard. *rows* maps names to Annual Model rows."""
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_properties.tabColor = DARK_TEAL

    # Column widths
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 18

    end_col = 8  # Column H

    # Helper: sum a row across all years in Annual Model (cols C-V)
    def _am_sum(am_row):
        parts = [f"'Annual Model'!{get_column_letter(2+y)}{am_row}"
                 for y in range(1, NUM_YEARS + 1)]
        return "=" + "+".join(parts)

    # Helper: comma-joined refs for AVERAGE/MAX/MIN
    def _am_refs(am_row):
        return ",".join(
            "'Annual Model'!" + get_column_letter(2 + y) + str(am_row)
            for y in range(1, NUM_YEARS + 1)
        )

    last_yr_cl = get_column_letter(2 + NUM_YEARS)

    # ── Title Row ──
    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    title_cell = ws.cell(row=row, column=1, value="Nuveen Housing LP Financial Model")
    title_cell.font = TITLE_FONT
    title_cell.fill = HEADER_FILL
    title_cell.alignment = CENTER
    for c in range(1, end_col + 1):
        ws.cell(row=row, column=c).fill = HEADER_FILL

    # ── Fund Parameters Section ──
    row = 3
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    ws.cell(row=row, column=1, value="Fund Parameters").font = HEADER_FONT
    style_header_row(ws, row, 1, end_col)

    params = [
        (5, "Initial AUM", DEFAULT_AUM, FMT_CURRENCY),
        (6, "Management Fee", DEFAULT_MGMT_FEE_BPS, FMT_BPS),
        (7, "Performance Fee (% of excess return)", DEFAULT_PERF_FEE_BPS, FMT_BPS),
        (8, "Hurdle Rate", DEFAULT_HURDLE_BPS, FMT_BPS),
        (9, "Other Expenses", DEFAULT_OTHER_EXP_BPS, FMT_BPS),
    ]
    for r, label, default, fmt in params:
        label_cell = ws.cell(row=r, column=1, value=label)
        style_label_cell(label_cell)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        inp = ws.cell(row=r, column=3, value=default)
        style_input_cell(inp, fmt)

    # ── Summary Statistics Section ──
    row = 13
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    ws.cell(row=row, column=1, value="Summary Statistics").font = HEADER_FONT
    style_header_row(ws, row, 1, end_col)

    net_yield_refs = _am_refs(rows["net_yield"])
    gross_yield_refs = _am_refs(rows["gross_yield"])
    end_aum_refs = _am_refs(rows["ending_aum"])

    # ── Return Metrics ──
    row = 14
    ws.cell(row=row, column=1, value="Return Metrics").font = SUBHEADER_FONT
    style_subheader_row(ws, row, 1, end_col)
    ws.cell(row=row, column=1).alignment = LEFT

    stats = [
        (15, "Net IRR (Avg Annual Net Yield)",
         f"=AVERAGE({net_yield_refs})", FMT_PCT),
        (16, "Gross IRR (Avg Annual Gross Yield)",
         f"=AVERAGE({gross_yield_refs})", FMT_PCT),
        (17, "MOIC (Multiple on Invested Capital)",
         f"='Annual Model'!{last_yr_cl}{rows['ending_aum']}/Summary!$C$5",
         '0.00x'),
        (18, "DPI (Distributions to Paid-In)",
         f"=({_am_sum(rows['net_income'])[1:]})/Summary!$C$5",
         '0.00x'),
        (19, "Total Value to Paid-In (TVPI)",
         f"=('Annual Model'!{last_yr_cl}{rows['ending_aum']}+"
         f"{_am_sum(rows['net_income'])[1:]})/Summary!$C$5",
         '0.00x'),
    ]

    for r, label, formula, fmt in stats:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        label_cell = ws.cell(row=r, column=1, value=label)
        style_label_cell(label_cell, indent=1)
        val_cell = ws.cell(row=r, column=3, value=formula)
        style_calc_cell(val_cell, fmt, bold=True)

    # ── Capital Metrics ──
    row = 21
    ws.cell(row=row, column=1, value="Capital Metrics").font = SUBHEADER_FONT
    style_subheader_row(ws, row, 1, end_col)
    ws.cell(row=row, column=1).alignment = LEFT

    cap_stats = [
        (22, "Initial Committed Capital",
         "=Summary!$C$5", FMT_CURRENCY),
        (23, "Ending AUM (Year 20)",
         f"='Annual Model'!{last_yr_cl}{rows['ending_aum']}", FMT_CURRENCY),
        (24, "Total AUM Growth",
         f"='Annual Model'!{last_yr_cl}{rows['ending_aum']}/Summary!$C$5-1", FMT_PCT),
        (25, "Total Cumulative Net Income",
         _am_sum(rows["net_income"]), FMT_CURRENCY),
        (26, "Total Cumulative Gross Income",
         _am_sum(rows["total_gross"]), FMT_CURRENCY),
        (27, "Peak AUM",
         f"=MAX({end_aum_refs})", FMT_CURRENCY),
    ]

    for r, label, formula, fmt in cap_stats:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        label_cell = ws.cell(row=r, column=1, value=label)
        style_label_cell(label_cell, indent=1)
        val_cell = ws.cell(row=r, column=3, value=formula)
        style_calc_cell(val_cell, fmt, bold=True)

    # ── Fee & Expense Metrics ──
    row = 29
    ws.cell(row=row, column=1, value="Fee & Expense Metrics").font = SUBHEADER_FONT
    style_subheader_row(ws, row, 1, end_col)
    ws.cell(row=row, column=1).alignment = LEFT

    fee_stats = [
        (30, "Total Management Fees",
         _am_sum(rows["mgmt_fee"]), FMT_CURRENCY),
        (31, "Total Performance Fees",
         _am_sum(rows["perf_fee"]), FMT_CURRENCY),
        (32, "Total Other Expenses",
         _am_sum(rows["other_exp"]), FMT_CURRENCY),
        (33, "Total All-In Fees & Expenses",
         _am_sum(rows["total_fees"]), FMT_CURRENCY),
        (34, "Fee Drag (Fees / Gross Income)",
         "=IF(C26=0,0,C33/C26)", FMT_PCT),
        (35, "Net-to-Gross Ratio",
         "=IF(C26=0,0,C25/C26)", FMT_PCT),
    ]

    for r, label, formula, fmt in fee_stats:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        label_cell = ws.cell(row=r, column=1, value=label)
        style_label_cell(label_cell, indent=1)
        val_cell = ws.cell(row=r, column=3, value=formula)
        style_calc_cell(val_cell, fmt, bold=True)

    # ── Yield & Risk Metrics ──
    row = 37
    ws.cell(row=row, column=1, value="Yield & Risk Metrics").font = SUBHEADER_FONT
    style_subheader_row(ws, row, 1, end_col)
    ws.cell(row=row, column=1).alignment = LEFT

    yield_stats = [
        (38, "Year 1 Net Yield",
         f"='Annual Model'!C{rows['net_yield']}", FMT_PCT),
        (39, f"Year {NUM_YEARS} Net Yield",
         f"='Annual Model'!{last_yr_cl}{rows['net_yield']}", FMT_PCT),
        (40, "Best Year Net Yield",
         f"=MAX({net_yield_refs})", FMT_PCT),
        (41, "Worst Year Net Yield",
         f"=MIN({net_yield_refs})", FMT_PCT),
        (42, "Year 1 Gross Yield",
         f"='Annual Model'!C{rows['gross_yield']}", FMT_PCT),
        (43, "Gross-to-Net Spread (Avg)",
         "=C16-C15", FMT_PCT),
    ]

    for r, label, formula, fmt in yield_stats:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        label_cell = ws.cell(row=r, column=1, value=label)
        style_label_cell(label_cell, indent=1)
        val_cell = ws.cell(row=r, column=3, value=formula)
        style_calc_cell(val_cell, fmt, bold=True)

    # ── Leverage & Securitization Metrics ──
    row = 45
    ws.cell(row=row, column=1,
            value="Leverage & Securitization Metrics").font = SUBHEADER_FONT
    style_subheader_row(ws, row, 1, end_col)
    ws.cell(row=row, column=1).alignment = LEFT

    lev_stats = [
        (46, "Total Leveraged AUM (Year 1)",
         f"='Annual Model'!C{rows['leveraged_amt']}", FMT_CURRENCY),
        (47, "Leverage Ratio (Year 1)",
         f"=IF('Annual Model'!C{rows['beg_aum']}=0,0,"
         f"'Annual Model'!C{rows['leveraged_amt']}/'Annual Model'!C{rows['beg_aum']})",
         FMT_PCT),
        (48, "Cumulative Net Leverage Income",
         _am_sum(rows["net_lev_inc"]), FMT_CURRENCY),
        (49, "Cumulative Net Securitization Income",
         _am_sum(rows["total_secur_inc"]), FMT_CURRENCY),
        (50, "Total Leverage + Securitization Income",
         "=C48+C49", FMT_CURRENCY),
    ]

    for r, label, formula, fmt in lev_stats:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        label_cell = ws.cell(row=r, column=1, value=label)
        style_label_cell(label_cell, indent=1)
        val_cell = ws.cell(row=r, column=3, value=formula)
        style_calc_cell(val_cell, fmt, bold=True)

    # ── Distributable Income Dynamics ──
    row = 52
    ws.cell(row=row, column=1,
            value="Distributable Income Dynamics").font = SUBHEADER_FONT
    style_subheader_row(ws, row, 1, end_col)
    ws.cell(row=row, column=1).alignment = LEFT

    # Column headers: Year 1 | Year 5 | Year 10 | Year 15 | Year 20 | Cumulative
    row = 53
    snapshot_years = [1, 5, 10, 15, NUM_YEARS]
    snapshot_labels = [f"Year {y}" for y in snapshot_years] + ["Cumulative"]
    for i, lbl in enumerate(snapshot_labels):
        hdr = ws.cell(row=row, column=3 + i, value=lbl)
        hdr.font = make_font(bold=True, color="FFFFFF", size=9)
        hdr.fill = HEADER_FILL
        hdr.alignment = CENTER
        hdr.border = THIN_BORDER

    def _snap_refs(am_row):
        """Excel refs for the five snapshot years."""
        return [f"='Annual Model'!{get_column_letter(2 + y)}{am_row}"
                for y in snapshot_years]

    def _snap_row(ws_row, label, am_row, fmt, cumulative_formula):
        ws.merge_cells(start_row=ws_row, start_column=1,
                       end_row=ws_row, end_column=2)
        lc = ws.cell(row=ws_row, column=1, value=label)
        style_label_cell(lc, indent=1)
        for col_offset, formula in enumerate(_snap_refs(am_row)):
            cell = ws.cell(row=ws_row, column=3 + col_offset, value=formula)
            style_calc_cell(cell, fmt)
        cum_cell = ws.cell(row=ws_row, column=8, value=cumulative_formula)
        style_calc_cell(cum_cell, fmt, bold=True)

    _snap_row(54, "Net Distributable Income",
              rows["net_income"], FMT_CURRENCY, _am_sum(rows["net_income"]))
    _snap_row(55, "Total Gross Income",
              rows["total_gross"], FMT_CURRENCY, _am_sum(rows["total_gross"]))
    _snap_row(56, "Asset Income",
              rows["total_asset_inc"], FMT_CURRENCY,
              _am_sum(rows["total_asset_inc"]))
    _snap_row(57, "Leverage Income",
              rows["net_lev_inc"], FMT_CURRENCY, _am_sum(rows["net_lev_inc"]))
    _snap_row(58, "Securitization Income",
              rows["total_secur_inc"], FMT_CURRENCY,
              _am_sum(rows["total_secur_inc"]))
    _snap_row(59, "Total Fees & Expenses",
              rows["total_fees"], FMT_CURRENCY, _am_sum(rows["total_fees"]))

    # Payout Ratio (Net / Gross) — snapshot + average
    ws.merge_cells(start_row=60, start_column=1, end_row=60, end_column=2)
    pr_label = ws.cell(row=60, column=1, value="Payout Ratio (Net / Gross)")
    style_label_cell(pr_label, indent=1)
    for col_offset, yr in enumerate(snapshot_years):
        yr_cl = get_column_letter(2 + yr)
        cell = ws.cell(
            row=60, column=3 + col_offset,
            value=f"=IFERROR('Annual Model'!{yr_cl}{rows['net_income']}/"
                  f"'Annual Model'!{yr_cl}{rows['total_gross']},0)")
        style_calc_cell(cell, FMT_PCT)
    avg_pr = ws.cell(
        row=60, column=8,
        value=f"=IFERROR(C54/C55,0)")
    style_calc_cell(avg_pr, FMT_PCT, bold=True)

    # ── Income Growth Metrics ──
    row = 62
    ws.cell(row=row, column=1,
            value="Income Growth Metrics").font = SUBHEADER_FONT
    style_subheader_row(ws, row, 1, end_col)
    ws.cell(row=row, column=1).alignment = LEFT

    ni_c1 = f"'Annual Model'!C{rows['net_income']}"
    ni_last = f"'Annual Model'!{last_yr_cl}{rows['net_income']}"
    net_income_refs = _am_refs(rows["net_income"])

    growth_stats = [
        (63, "Income CAGR (Year 1 → Year 20)",
         f"=IFERROR(({ni_last}/{ni_c1})^(1/{NUM_YEARS - 1})-1,0)",
         FMT_PCT),
        (64, "Average Annual Net Income",
         f"=C25/{NUM_YEARS}", FMT_CURRENCY),
        (65, "Best Year Net Income",
         f"=MAX({net_income_refs})", FMT_CURRENCY),
        (66, "Worst Year Net Income",
         f"=MIN({net_income_refs})", FMT_CURRENCY),
        (67, "Year-1 to Year-20 Net Income Growth",
         f"=IFERROR({ni_last}/{ni_c1}-1,0)", FMT_PCT),
    ]

    for r, label, formula, fmt in growth_stats:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        lc = ws.cell(row=r, column=1, value=label)
        style_label_cell(lc, indent=1)
        vc = ws.cell(row=r, column=3, value=formula)
        style_calc_cell(vc, fmt, bold=True)

    # ── Legend ──
    row = 71
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    legend = ws.cell(row=row, column=1,
                     value="Input cells are highlighted in orange. "
                           "Modify inputs to update the model.")
    legend.font = make_font(italic=True, color=MEDIUM_GRAY, size=9)

    swatch = ws.cell(row=row, column=5)
    swatch.fill = INPUT_FILL
    swatch.value = "= Input"
    swatch.font = make_font(size=9, color="1A237E")
    swatch.border = THIN_BORDER

    return ws


# =============================================================================
# MAIN
# =============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="Generate Nuveen Housing LP Financial Model")
    parser.add_argument(
        "--bloomberg", action="store_true",
        help="Pull SOFR/SIFMA rates from Bloomberg Terminal (requires blpapi)")
    args = parser.parse_args()

    # Optionally pull Bloomberg rates
    sofr_curve = None
    sifma_curve = None
    if args.bloomberg:
        bbg = fetch_bloomberg_rates()
        if bbg:
            sofr_curve = bbg.get("sofr")
            sifma_curve = bbg.get("sifma")

    wb = Workbook()

    # Build Annual Model first (need row mappings for Summary)
    am_ws, rows = build_annual_model_sheet(wb)

    # Build Summary (uses row mappings)
    build_summary_sheet(wb, rows)

    # Build Market Rates sheet (inserted between Summary and Annual Model)
    build_market_rates_sheet(wb, sofr_curve=sofr_curve, sifma_curve=sifma_curve)

    # Reorder sheets: Summary, Market Rates, Annual Model
    wb._sheets = [wb["Summary"], wb["Market Rates"], wb["Annual Model"]]

    output_file = "LP_Financial_Model.xlsx"
    wb.save(output_file)
    print(f"Financial model generated: {output_file}")


if __name__ == "__main__":
    main()
